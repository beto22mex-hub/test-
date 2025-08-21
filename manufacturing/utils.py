from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import io


class ExportUtils:
    """Utility class for exporting data to Excel and PDF"""
    
    @staticmethod
    def export_to_excel(queryset):
        """Export SerialNumber queryset to Excel"""
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Números de Serie"
        
        # Define headers
        headers = [
            'Número de Serie', 'Número de Orden', 'Componente', 'Descripción',
            'Estado', 'Progreso (%)', 'Creado Por', 'Fecha Creación', 'Fecha Completado'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Write data
        for row, serial in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=serial.serial_number)
            ws.cell(row=row, column=2, value=serial.order_number)
            ws.cell(row=row, column=3, value=serial.authorized_part.part_number)
            ws.cell(row=row, column=4, value=serial.authorized_part.description)
            ws.cell(row=row, column=5, value=serial.get_status_display())
            ws.cell(row=row, column=6, value=f"{serial.completion_percentage}%")
            ws.cell(row=row, column=7, value=serial.created_by.get_full_name())
            ws.cell(row=row, column=8, value=serial.created_at.strftime('%Y-%m-%d %H:%M'))
            ws.cell(row=row, column=9, value=serial.completed_at.strftime('%Y-%m-%d %H:%M') if serial.completed_at else '')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="numeros_serie_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        
        wb.save(response)
        return response
    
    @staticmethod
    def export_to_pdf(queryset):
        """Export SerialNumber queryset to PDF"""
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        # Title
        title = Paragraph("Reporte de Números de Serie", title_style)
        elements.append(title)
        elements.append(Spacer(1, 20))
        
        # Date
        date_text = f"Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
        date_para = Paragraph(date_text, styles['Normal'])
        elements.append(date_para)
        elements.append(Spacer(1, 20))
        
        # Table data
        data = [['Número de Serie', 'Orden', 'Componente', 'Estado', 'Progreso']]
        
        for serial in queryset:
            data.append([
                serial.serial_number,
                serial.order_number,
                serial.authorized_part.part_number,
                serial.get_status_display(),
                f"{serial.completion_percentage}%"
            ])
        
        # Create table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        
        # Return response
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="numeros_serie_{timezone.now().strftime("%Y%m%d_%H%M")}.pdf"'
        
        return response


class NotificationService:
    """Service for sending real-time notifications"""
    
    @staticmethod
    def send_process_notification(serial_number, operation, status, user):
        """Send process update notification via WebSocket"""
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to notifications group
        async_to_sync(channel_layer.group_send)(
            "notifications",
            {
                "type": "process_update",
                "serial_number": serial_number,
                "operation": operation,
                "status": status,
                "user": user,
                "timestamp": str(timezone.now())
            }
        )
        
        # Send to dashboard group
        async_to_sync(channel_layer.group_send)(
            "dashboard",
            {
                "type": "dashboard_update",
                "data": {
                    "type": "process_update",
                    "serial_number": serial_number,
                    "timestamp": str(timezone.now())
                }
            }
        )
    
    @staticmethod
    def send_alert_notification(alert_type, message, priority="MEDIUM"):
        """Send general alert notification"""
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        async_to_sync(channel_layer.group_send)(
            "notifications",
            {
                "type": "notification_message",
                "message": message,
                "alert_type": alert_type,
                "priority": priority,
                "timestamp": str(timezone.now())
            }
        )
